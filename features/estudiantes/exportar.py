import os
from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

from features.estudiantes.models import Estudiante
from features.profesores.models import Profesor

# ── Paleta institucional ───────────────────────────────────────────────────────
COLOR_PRIMARIO   = "1A3A5C"  # Azul UPEL oscuro
COLOR_SECUNDARIO = "2E86C1"  # Azul medio — encabezado tabla
COLOR_FILA_PAR   = "EBF5FB"  # Azul muy claro
COLOR_ACENTO     = "F39C12"  # Naranja — totales

LOGO_PATH = os.path.join(
    os.path.dirname(__file__), "..", "..", "recursos", "icons", "encabezado_excel.png"
)

# ── Helpers de estilo ─────────────────────────────────────────────────────────
def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=10, color="000000", name="Calibri") -> Font:
    return Font(bold=bold, size=size, color=color, name=name)

def _border_thin() -> Border:
    s = Side(style="thin", color="BDC3C7")
    return Border(left=s, right=s, top=s, bottom=s)

def _center(wrap=False) -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _left(wrap=False) -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


# ── Encabezado institucional ──────────────────────────────────────────────────
def _escribir_encabezado(ws, subtitulo: str, fecha_gen: str) -> None:
    n_cols = len(COLUMNAS)

    # Alturas de fila — rows 1+2 = 90pt = 120px (altura natural de la imagen)
    ws.row_dimensions[1].height = 45   # imagen encabezado (fila superior)
    ws.row_dimensions[2].height = 45   # imagen encabezado (fila inferior)
    ws.row_dimensions[3].height = 4    # separador
    ws.row_dimensions[4].height = 24   # subtítulo
    ws.row_dimensions[5].height = 16   # fecha
    ws.row_dimensions[6].height = 6    # espacio
    ws.row_dimensions[7].height = 4    # separador antes de tabla

    # Imagen del encabezado institucional — ancho completo de la hoja
    logo_abs = os.path.normpath(LOGO_PATH)
    if os.path.isfile(logo_abs):
        img        = XLImage(logo_abs)
        img.width  = 4000
        img.height = 120
        img.anchor = "A1"
        ws.add_image(img)

    # Separador horizontal (fila 3)
    for col in range(1, n_cols + 1):
        ws.cell(row=3, column=col).fill = _fill(COLOR_SECUNDARIO)

    # Fila 4 — subtítulo del documento
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=n_cols)
    sub = ws.cell(row=4, column=1, value=subtitulo)
    sub.font      = _font(bold=True, size=13, color="1A3A5C")
    sub.alignment = Alignment(horizontal="center", vertical="center")

    # Fila 5 — fecha de generación
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=n_cols)
    fecha_cell = ws.cell(row=5, column=1, value=f"Generado el {fecha_gen}")
    fecha_cell.font      = _font(size=9, color="7F8C8D")
    fecha_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Separador antes de la tabla (fila 7)
    for col in range(1, n_cols + 1):
        ws.cell(row=7, column=col).fill = _fill(COLOR_PRIMARIO)


# ── Encabezado de tabla ───────────────────────────────────────────────────────
COLUMNAS = [
    # (header, ancho, clave_dato)
    ("Tipo Doc.",          11, "tipo_doc"),
    ("Cédula / Pasaporte", 18, "cedula"),
    ("Nombre",             18, "nombre"),
    ("Apellido",           18, "apellido"),
    ("Correo",             28, "correo"),
    ("Cohorte",            10, "cohorte"),
    ("Maestría",           28, "maestria"),
    ("Título del Proyecto",40, "titulo"),
    ("Solvencia",          14, "solvencia"),
    ("Estado",             14, "estado"),
    ("Tutor Principal",    26, "tutor_p"),
    ("Tutor Suplente",     26, "tutor_s"),
    ("Jurado 1 Principal", 26, "j1_p"),
    ("Jurado 1 Suplente",  26, "j1_s"),
    ("Jurado 2 Principal", 26, "j2_p"),
    ("Jurado 2 Suplente",  26, "j2_s"),
    ("Recibo M1",          16, "recibo_m1"),
    ("Monto M1",           12, "monto_m1"),
    ("Fecha Trans. M1",    16, "fecha_m1"),
    ("Recibo Caja M1",     14, "recibo_caja_m1"),
    ("Recibo M2",          16, "recibo_m2"),
    ("Monto M2",           12, "monto_m2"),
    ("Fecha Trans. M2",    16, "fecha_m2"),
    ("Recibo Caja M2",     14, "recibo_caja_m2"),
    ("Recibo M3",          16, "recibo_m3"),
    ("Monto M3",           12, "monto_m3"),
    ("Fecha Trans. M3",    16, "fecha_m3"),
    ("Recibo Caja M3",     14, "recibo_caja_m3"),
]

FILA_ENCABEZADO_TABLA = 8


def _escribir_encabezado_tabla(ws) -> None:
    for col_idx, (header, ancho, _) in enumerate(COLUMNAS, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = ancho

        cell = ws.cell(row=FILA_ENCABEZADO_TABLA, column=col_idx, value=header)
        cell.font      = _font(bold=True, size=10, color="FFFFFF")
        cell.fill      = _fill(COLOR_SECUNDARIO)
        cell.alignment = _center(wrap=True)
        cell.border    = _border_thin()
    ws.row_dimensions[FILA_ENCABEZADO_TABLA].height = 30


# ── Resolución de IDs a nombres ───────────────────────────────────────────────
def _nombre_prof(prof_map: dict[int, str], id_prof: int) -> str:
    return prof_map.get(id_prof, "—") if id_prof else "—"


def _badge_solvencia(est: Estudiante) -> str:
    if not est.verificado_m1:
        return "Pendiente M1"
    if not est.verificado_m2:
        return "Pendiente M2"
    if not est.verificado_m3:
        return "Solvente M2"
    return "Solvente M3"


def _fmt_fecha(d) -> str:
    if not d:
        return "—"
    if isinstance(d, date):
        return d.strftime("%d/%m/%Y")
    return str(d)


def _fmt_monto(v) -> str:
    if v is None:
        return "—"
    try:
        return f"Bs. {float(v):.2f}"
    except (ValueError, TypeError):
        return "—"


# ── Escritura de filas de datos ───────────────────────────────────────────────
def _escribir_filas(
    ws,
    estudiantes: list[Estudiante],
    maestria_map: dict[int, str],
    prof_map: dict[int, str],
    ids_asignados: set[int],
) -> None:
    for row_offset, est in enumerate(estudiantes):
        fila = FILA_ENCABEZADO_TABLA + 1 + row_offset
        es_par = row_offset % 2 == 1
        bg = COLOR_FILA_PAR if es_par else "FFFFFF"

        tipo_doc = "Cédula" if est.tipo_documento == "Cedula" else est.tipo_documento
        doc_num  = est.cedula
        if est.pasaporte:
            doc_num = f"{est.cedula} / {est.pasaporte}"

        datos = {
            "tipo_doc":       tipo_doc,
            "cedula":         doc_num,
            "nombre":         est.nombre,
            "apellido":       est.apellido,
            "correo":         est.correo_electronico or "—",
            "cohorte":        est.cohorte or "—",
            "maestria":       maestria_map.get(est.maestria_id, "—"),
            "titulo":         est.titulo_proyecto or "—",
            "solvencia":      _badge_solvencia(est),
            "estado":         "Asignado" if est.id in ids_asignados else "Disponible",
            "tutor_p":        _nombre_prof(prof_map, est.id_tutor_principal),
            "tutor_s":        _nombre_prof(prof_map, est.id_tutor_suplente),
            "j1_p":           _nombre_prof(prof_map, est.id_jurado1_principal),
            "j1_s":           _nombre_prof(prof_map, est.id_jurado1_suplente),
            "j2_p":           _nombre_prof(prof_map, est.id_jurado2_principal),
            "j2_s":           _nombre_prof(prof_map, est.id_jurado2_suplente),
            "recibo_m1":      est.recibo_m1 or "—",
            "monto_m1":       _fmt_monto(est.monto_m1),
            "fecha_m1":       _fmt_fecha(est.dia_transferencia_m1),
            "recibo_caja_m1": est.recibo_caja_m1 or "—",
            "recibo_m2":      est.recibo_m2 or "—",
            "monto_m2":       _fmt_monto(est.monto_m2),
            "fecha_m2":       _fmt_fecha(est.dia_transferencia_m2),
            "recibo_caja_m2": est.recibo_caja_m2 or "—",
            "recibo_m3":      est.recibo_m3 or "—",
            "monto_m3":       _fmt_monto(est.monto_m3),
            "fecha_m3":       _fmt_fecha(est.dia_transferencia_m3),
            "recibo_caja_m3": est.recibo_caja_m3 or "—",
        }

        for col_idx, (_, _, clave) in enumerate(COLUMNAS, start=1):
            cell = ws.cell(row=fila, column=col_idx, value=datos[clave])
            cell.fill      = _fill(bg)
            cell.font      = _font(size=9)
            cell.alignment = _left(wrap=False)
            cell.border    = _border_thin()

        ws.row_dimensions[fila].height = 16


# ── Fila de total ─────────────────────────────────────────────────────────────
def _escribir_total(ws, n_estudiantes: int) -> None:
    fila_total = FILA_ENCABEZADO_TABLA + 1 + n_estudiantes
    cell = ws.cell(row=fila_total, column=1,
                   value=f"Total de estudiantes: {n_estudiantes}")
    cell.font      = _font(bold=True, size=10, color="FFFFFF")
    cell.fill      = _fill(COLOR_ACENTO)
    cell.alignment = _left()
    cell.border    = _border_thin()
    ws.merge_cells(
        start_row=fila_total, start_column=1,
        end_row=fila_total,   end_column=len(COLUMNAS),
    )
    ws.row_dimensions[fila_total].height = 18


# ── Hoja de estadísticas ──────────────────────────────────────────────────────
def _tabla_stats(
    ws, fila: int, titulo: str, header: tuple, filas_data: list
) -> tuple[int, int, int]:
    """Escribe una tabla con cuadro estilizado. Devuelve (fila_header, inicio_data, fin_data)."""
    # Fila título
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=2)
    tc = ws.cell(row=fila, column=1, value=titulo)
    tc.font      = _font(bold=True, size=11, color="FFFFFF")
    tc.fill      = _fill(COLOR_PRIMARIO)
    tc.alignment = _center()
    tc.border    = _border_thin()
    ws.row_dimensions[fila].height = 18
    fila += 1

    # Fila header
    fila_header = fila
    for col_idx, texto in enumerate(header, start=1):
        cell = ws.cell(row=fila, column=col_idx, value=texto)
        cell.font      = _font(bold=True, size=10, color="FFFFFF")
        cell.fill      = _fill(COLOR_SECUNDARIO)
        cell.alignment = _center()
        cell.border    = _border_thin()
    ws.row_dimensions[fila].height = 16
    fila += 1

    inicio = fila
    for row_offset, (etiqueta, valor) in enumerate(filas_data):
        bg = COLOR_FILA_PAR if row_offset % 2 == 0 else "FFFFFF"
        c1 = ws.cell(row=fila, column=1, value=etiqueta)
        c2 = ws.cell(row=fila, column=2, value=valor)
        for cell in (c1, c2):
            cell.fill   = _fill(bg)
            cell.border = _border_thin()
            cell.font   = _font(size=9)
        c1.alignment = _left()
        c2.alignment = _center()
        ws.row_dimensions[fila].height = 15
        fila += 1

    return fila_header, inicio, fila - 1


def _dLbls(show_percent: bool = False) -> DataLabelList:
    d = DataLabelList()
    d.showVal        = True
    d.showPercent    = show_percent
    d.showLegendKey  = False
    d.showCatName    = show_percent   # en torta muestra la categoría
    d.showSerName    = False
    return d


def _agregar_estadisticas(
    wb: Workbook,
    estudiantes: list[Estudiante],
    maestria_map: dict[int, str],
    ids_asignados: set[int],
) -> None:
    ws = wb.create_sheet("Estadísticas")

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 4   # espaciador visual

    # ── Datos ──
    conteo_maestria: dict[str, int] = {}
    for est in estudiantes:
        nombre = maestria_map.get(est.maestria_id, "Sin maestría")
        conteo_maestria[nombre] = conteo_maestria.get(nombre, 0) + 1

    conteo_cohorte: dict[str, int] = {}
    for est in estudiantes:
        c = est.cohorte or "Sin cohorte"
        conteo_cohorte[c] = conteo_cohorte.get(c, 0) + 1

    asignados   = sum(1 for e in estudiantes if e.id in ids_asignados)
    disponibles = len(estudiantes) - asignados

    # ── Tablas con cuadro ──
    fila = 1
    h_m, ini_m, fin_m = _tabla_stats(
        ws, fila, "Estudiantes por Maestría", ("Maestría", "Nº"),
        sorted(conteo_maestria.items()),
    )
    fila = fin_m + 3

    h_c, ini_c, fin_c = _tabla_stats(
        ws, fila, "Estudiantes por Cohorte", ("Cohorte", "Nº"),
        sorted(conteo_cohorte.items()),
    )
    fila = fin_c + 3

    h_e, ini_e, fin_e = _tabla_stats(
        ws, fila, "Estado de Estudiantes", ("Estado", "Nº"),
        [("Asignados", asignados), ("Disponibles", disponibles)],
    )

    # ── Gráfico 1: barras por maestría (E1) ──
    chart1 = BarChart()
    chart1.type          = "col"
    chart1.title         = "Estudiantes por Maestría"
    chart1.y_axis.title  = "Cantidad"
    chart1.style         = 10
    chart1.width         = 22
    chart1.height        = 14
    chart1.dLbls         = _dLbls()
    chart1.add_data(Reference(ws, min_col=2, min_row=h_m, max_row=fin_m),
                    titles_from_data=True)
    chart1.set_categories(Reference(ws, min_col=1, min_row=ini_m, max_row=fin_m))
    ws.add_chart(chart1, "E1")

    # ── Gráfico 2: barras por cohorte (E34) ──
    chart2 = BarChart()
    chart2.type          = "col"
    chart2.title         = "Estudiantes por Cohorte"
    chart2.y_axis.title  = "Cantidad"
    chart2.style         = 10
    chart2.width         = 22
    chart2.height        = 14
    chart2.dLbls         = _dLbls()
    chart2.add_data(Reference(ws, min_col=2, min_row=h_c, max_row=fin_c),
                    titles_from_data=True)
    chart2.set_categories(Reference(ws, min_col=1, min_row=ini_c, max_row=fin_c))
    ws.add_chart(chart2, "E34")

    # ── Gráfico 3: torta estado (E67) ──
    chart3 = PieChart()
    chart3.title   = "Estado: Asignados vs Disponibles"
    chart3.style   = 10
    chart3.width   = 16
    chart3.height  = 14
    chart3.dLbls   = _dLbls(show_percent=True)
    chart3.add_data(Reference(ws, min_col=2, min_row=h_e, max_row=fin_e),
                    titles_from_data=True)
    chart3.set_categories(Reference(ws, min_col=1, min_row=ini_e, max_row=fin_e))
    ws.add_chart(chart3, "E67")


# ── Función pública ───────────────────────────────────────────────────────────
def generar_excel_estudiantes(
    estudiantes: list[Estudiante],
    profesores: list[Profesor],
    maestrias: list,
    ids_asignados: set[int],
    nombre_maestria: str | None = None,
) -> BytesIO:
    """
    Genera el Excel de registros de estudiantes.
    - nombre_maestria=None → todos + hoja de estadísticas con gráficos.
    - nombre_maestria="X"  → solo esa maestría, sin gráficos.
    Devuelve un BytesIO listo para StreamingResponse.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Estudiantes"

    maestria_map = {m.id: m.nombre for m in maestrias}
    prof_map     = {p.id: p.nombre_completo for p in profesores}

    subtitulo = f"Maestría: {nombre_maestria}" if nombre_maestria else "Registro General de Estudiantes"
    fecha_gen = date.today().strftime("%d/%m/%Y")

    # Anclar columnas A-B para el logo antes del encabezado de texto
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 10

    _escribir_encabezado(ws, subtitulo, fecha_gen)
    _escribir_encabezado_tabla(ws)
    _escribir_filas(ws, estudiantes, maestria_map, prof_map, ids_asignados)
    _escribir_total(ws, len(estudiantes))

    # Congelar filas de encabezado hasta la fila del encabezado de tabla
    ws.freeze_panes = f"A{FILA_ENCABEZADO_TABLA + 1}"

    if not nombre_maestria:
        _agregar_estadisticas(wb, estudiantes, maestria_map, ids_asignados)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
