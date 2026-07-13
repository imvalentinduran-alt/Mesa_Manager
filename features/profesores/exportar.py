import os
from io import BytesIO

from openpyxl import Workbook
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from features.profesores.models import Profesor

COLOR_PRIMARIO   = "1A3A5C"
COLOR_SECUNDARIO = "2E86C1"
COLOR_FILA_PAR   = "EBF5FB"

LOGO_PATH = os.path.join(
    os.path.dirname(__file__), "..", "..", "recursos", "icons", "encabezado_excel_profe.png"
)

COLUMNAS = [
    ("Cédula / Pasaporte", 20, "cedula"),
    ("Nombre",             22, "nombre"),
    ("Apellido",           22, "apellido"),
    ("Especialidad",       36, "especialidad"),
    ("Correo Electrónico", 36, "correo_electronico"),
    ("Disponible",         14, "is_active"),
]

FILA_ENCABEZADO_TABLA = 8


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=10, color="000000", name="Calibri") -> Font:
    return Font(bold=bold, size=size, color=color, name=name)

def _border_thin() -> Border:
    s = Side(style="thin", color="BDC3C7")
    return Border(left=s, right=s, top=s, bottom=s)


def _escribir_encabezado(ws, fecha_gen: str) -> None:
    n_cols = len(COLUMNAS)

    ws.row_dimensions[1].height = 45
    ws.row_dimensions[2].height = 45
    ws.row_dimensions[3].height = 4
    ws.row_dimensions[4].height = 24
    ws.row_dimensions[5].height = 16
    ws.row_dimensions[6].height = 6
    ws.row_dimensions[7].height = 4

    logo_abs = os.path.normpath(LOGO_PATH)
    if os.path.isfile(logo_abs):
        img        = XLImage(logo_abs)
        img.width  = 1120
        img.height = 120
        img.anchor = "A1"
        ws.add_image(img)

    for col in range(1, n_cols + 1):
        ws.cell(row=3, column=col).fill = _fill(COLOR_SECUNDARIO)

    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=n_cols)
    sub = ws.cell(row=4, column=1, value="Registro General de Profesores")
    sub.font      = _font(bold=True, size=13, color="1A3A5C")
    sub.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=n_cols)
    fecha_cell = ws.cell(row=5, column=1, value=f"Generado el {fecha_gen}")
    fecha_cell.font      = _font(size=9, color="7F8C8D")
    fecha_cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in range(1, n_cols + 1):
        ws.cell(row=7, column=col).fill = _fill(COLOR_PRIMARIO)


def _escribir_encabezado_tabla(ws) -> None:
    for col_idx, (header, ancho, _) in enumerate(COLUMNAS, start=1):
        cell = ws.cell(row=FILA_ENCABEZADO_TABLA, column=col_idx, value=header)
        cell.font      = _font(bold=True, size=10, color="FFFFFF")
        cell.fill      = _fill(COLOR_SECUNDARIO)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _border_thin()
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho
    ws.row_dimensions[FILA_ENCABEZADO_TABLA].height = 30


def _escribir_fila(ws, row_offset: int, profesor: Profesor) -> None:
    fila = FILA_ENCABEZADO_TABLA + 1 + row_offset
    es_par = row_offset % 2 == 0
    fill   = _fill(COLOR_FILA_PAR) if es_par else PatternFill()

    valores = [
        profesor.cedula,
        profesor.nombre,
        profesor.apellido,
        profesor.especialidad or "",
        profesor.correo_electronico or "",
        "Sí" if profesor.is_active else "No",
    ]

    for col_idx, valor in enumerate(valores, start=1):
        cell = ws.cell(row=fila, column=col_idx, value=valor)
        cell.font      = _font(size=10)
        cell.fill      = fill
        cell.border    = _border_thin()
        cell.alignment = Alignment(
            horizontal="center" if col_idx in (1, 6) else "left",
            vertical="center",
        )


def generar_excel_profesores(profesores: list[Profesor]) -> BytesIO:
    from datetime import date
    wb = Workbook()
    ws = wb.active
    ws.title = "Profesores"

    fecha_gen = date.today().strftime("%d/%m/%Y")
    _escribir_encabezado(ws, fecha_gen)
    _escribir_encabezado_tabla(ws)

    for i, prof in enumerate(profesores):
        _escribir_fila(ws, i, prof)

    # Fila total
    n = len(profesores)
    fila_total = FILA_ENCABEZADO_TABLA + 1 + n
    ws.merge_cells(start_row=fila_total, start_column=1, end_row=fila_total, end_column=len(COLUMNAS))
    total_cell = ws.cell(row=fila_total, column=1, value=f"Total: {n} profesor{'es' if n != 1 else ''}")
    total_cell.font      = _font(bold=True, size=10, color="FFFFFF")
    total_cell.fill      = _fill(COLOR_PRIMARIO)
    total_cell.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[fila_total].height = 22

    ws.freeze_panes = f"A{FILA_ENCABEZADO_TABLA + 1}"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
